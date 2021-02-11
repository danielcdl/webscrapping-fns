import json
import requests
from datetime import date
from time import sleep

from tqdm import tqdm
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side, BORDER_THIN


colunas = ('A', 'B', 'C', 'D', 'E', 'F', 'G','H','I','J','K','L','M','N','O')
meses = ('Jan', 'Fev', 'Mar', 'Abr', 'mai', 'jun', 'jul', 'ago', 'set', 'out', 'nov', 'dez', 'Valor Total')
nome_tabela = ('ATENCAO BÁSICA', 'VIGILÂNCIA EM SAÚDE', 'ASSISTÊNCIA FARMACÊUTICA')

nome_tabela_covid = ('COMPETÊNCIA/ PARCELA', 'Nº OB', 'DATA OB', 'TIPO DE REPASSE',	'BANCO OB',	'AGÊNCIA OB',
    'CONTA OB', 'VALOR TOTAL', 'DESCONTO', 'LÍQUIDO', 'REJEIÇÃO', 'PROCESSO', 'Nº PROPOSTA', 'PORTARIA')

municipios = (
    # {'nome': 'Anajás', 'cpfCnpjUg': 13715424000184, 'municipio': 150070},
    # {'nome': 'Bagre', 'cpfCnpjUg': 13888332000104, 'municipio': 150110},
    {'nome': 'Breves', 'cpfCnpjUg': 17298800000133, 'municipio': 150180},
    # {'nome': 'Curralinho', 'cpfCnpjUg': 11441240000148, 'municipio': 150280},
    # {'nome': 'Gurupá', 'cpfCnpjUg': 12049775000130, 'municipio': 150310},
    # {'nome': 'Melgaço', 'cpfCnpjUg': 11530230000189, 'municipio': 150450},
    # {'nome': 'Portel', 'cpfCnpjUg': 11956268000118, 'municipio': 1}
)

URL_ACAO = 'https://consultafns.saude.gov.br/recursos/consulta-detalhada/detalhe-acao'
URL_PAGAMENTO = 'https://consultafns.saude.gov.br/recursos/consulta-detalhada/detalhe-pagamento'

grupos = (12, 34, 35)

while True:
    tipo = input('Digite 1 para o ano atual ou 2 para anos anteriores\n')
    if tipo == '1':
        data = date.today()
        mes_atual = data.month
        ano = data.year
        break
    elif tipo == '2':
        mes_atual = 12
        try:
            ano = int(input('Digite o ano a ser pesquisado: '))
            if ano > 2008:
                break
            else:
                print('Digite uma data mais atual')
        except ValueError:
            print('Digite uma data válida')


dados_geral = {
    'ano': ano, 
    'blocos': 10,
    'count': 100,
    'estado': 'PA',
    'page': 1,
    'tipoConsulta': 2
}


def lista_chaves(parametros):
    chaves = {}
    descricoes = []
    while True:
        try:
            req = requests.get(URL_ACAO, params=parametros)
            if req.status_code == 200:
                i = 0
                resposta = json.loads(req.content)['resultado']['dados']
                for dado in resposta:
                    chave = dado['id']
                    if chave not in chaves:
                        chaves[chave] = i
                        descricoes.append(dado['descricao'])
                        i += 1
                break
            else:
                print('\nfalha', req.status_code)
        except requests.exceptions.ConnectionError:
            print('=' * 15)
            print('\nErro na conecção! tentando novamente...')
            print('=' * 15)
            sleep(1)
    return chaves, descricoes

def dados_tabela(parametros):
    tabela = {}
    while True:
        try:  
            req = requests.get(URL_ACAO, params=parametros)
            if req.status_code == 200:
                resposta = json.loads(req.content)['resultado']['dados']
                for dado in resposta:
                    chave = dado['id']
                    if chave != 0:
                        valor = dado['valorLiquido']
                        tabela[chave] = valor
                break
            else:
                print('=' * 15)
                print('\nfalha', req.status_code)
                print('=' * 15)
        except requests.exceptions.ConnectionError:
            print('=' * 15)
            print('\nErro na conecção! tentando novamente...')
            print('=' * 15)
            sleep(1)
    return tabela

def covid_tabela(parametros):
    tabela = []
    while True:
        try:  
            req = requests.get(URL_PAGAMENTO, params=parametros)
            if req.status_code == 200:
                resposta = json.loads(req.content)['resultado']['dados']
                for acao in resposta:
                    corona = (
                        acao['competencia'],
                        acao['numeroDocumentoSiafi'],
                        acao['dataCriacaoSiafi'],
                        acao['id']['esferaAdministrativa'],
                        acao['codigoBanco'],
                        acao['codigoAgencia'],
                        acao['contaCorrente'],
                        acao['valorTotal'],
                        acao['valorDescontoTotal'],
                        acao['valorLiquido'],
                        acao['motivoRejeicao'],
                        acao['id']['processoFormatado'],
                        '',
                        acao['nuPortaria']
                    )
                    tabela.append(corona)
                break
            else:
                print('=' * 15)
                print('falha', req.status_code)
                print('=' * 15)
        except requests.exceptions.ConnectionError:
            print('=' * 15)
            print('Erro na conecção! tentando novamente...')
            print('=' * 15)
            sleep(1)
    return tabela


try:
    planilha = load_workbook(filename=f'Planilha de acompanhamento {ano}.xlsx')
except FileNotFoundError:
    planilha = load_workbook(filename='modelo/modelo.xlsx')

bordas = Border(
    left=Side(border_style=BORDER_THIN, color='00000000'),
    right=Side(border_style=BORDER_THIN, color='00000000'),
    top=Side(border_style=BORDER_THIN, color='00000000'),
    bottom=Side(border_style=BORDER_THIN, color='00000000')
)

pbar = tqdm(total=len(municipios) * 3 * mes_atual + 1, desc='Iniciando')
for municipio in municipios:
    pbar.desc = f"Baixando {municipio['nome']}"
    dados = dados_geral
    dados['cpfCnpjUg'] = municipio['cpfCnpjUg']
    dados['municipio'] = municipio['municipio']
    
    tabela_geral = []
    descricoes_geral = []
    for grupo in grupos:  
        dados['grupo'] = grupo
        chaves, descricoes = lista_chaves(dados)
        descricoes_geral.append(descricoes)
        tabela = []
        for i in range(len(chaves)):
            tabela.append(['']*12)
        for mes in range(1, mes_atual + 1):
            dados['mes'] = mes
            valores = dados_tabela(dados)

            for chave in valores:
                tabela[chaves[chave]][mes-1] = valores[chave]
            pbar.update(1)
        del dados['mes']
        tabela_geral.append(tabela)

    dados['grupo'] = 245
    dados['componentes'] = 175
    chaves = lista_chaves(dados)
    tabela = []
    del dados['grupo']
    for chave in chaves[0]:
        dados['acoes'] = chave
        valores = covid_tabela(dados)
        for valor in valores:
            tabela.append(valor)
    pbar.update(1)
    tabela_geral.append(tabela)
    del dados['acoes'] 
    del dados['componentes']

    #=============== Salvar Planilha ===========================

    aba = planilha[municipio['nome']]

    linha = 14
    for linha_grupo, descricao in zip(tabela_geral[0], descricoes_geral[0]):
        aba[f'{colunas[1]}{linha}'] = descricao
        aba[f'{colunas[1]}{linha}'].border = bordas
        aba[f'{colunas[14]}{linha}'] = f'=SUM({colunas[2]}{linha}:{colunas[13]}{linha})'
        aba[f'{colunas[14]}{linha}'].font = Font(bold=True)
        aba[f'{colunas[14]}{linha}'].border = bordas
        for k in range(12):
            aba[f'{colunas[k + 2]}{linha}'] = linha_grupo[k]
            aba[f'{colunas[k + 2]}{linha}'].border = bordas
        linha += 1  
    aba[f'{colunas[1]}{linha}'] = 'Subtotal Componente'
    aba[f'{colunas[1]}{linha}'].border = bordas
    aba[f'{colunas[1]}{linha}'].font = Font(bold=True)
    for k in range(13):
        aba[f'{colunas[k + 2]}{linha}'] = f'=SUM({colunas[k + 2]}14:{colunas[k + 2]}{linha-1})'
        aba[f'{colunas[k + 2]}{linha}'].border = bordas
        aba[f'{colunas[k + 2]}{linha}'].font = Font(bold=True)

    
    linha = 28
    for linha_grupo, descricao in zip(tabela_geral[1], descricoes_geral[1]):
        aba[f'{colunas[1]}{linha}'] = descricao
        aba[f'{colunas[1]}{linha}'].border = bordas
        aba[f'{colunas[14]}{linha}'] = f'=SUM({colunas[2]}{linha}:{colunas[13]}{linha})'
        aba[f'{colunas[14]}{linha}'].font = Font(bold=True)
        aba[f'{colunas[14]}{linha}'].border = bordas
        for k in range(12):
            aba[f'{colunas[k + 2]}{linha}'] = linha_grupo[k]
            aba[f'{colunas[k + 2]}{linha}'].border = bordas
        linha += 1  
    aba[f'{colunas[1]}{linha}'] = 'Subtotal Componente'
    aba[f'{colunas[1]}{linha}'].border = bordas
    for k in range(13):
        aba[f'{colunas[k + 2]}{linha}'] = f'=SUM({colunas[k + 2]}28:{colunas[k + 2]}{linha-1})'
        aba[f'{colunas[k + 2]}{linha}'].border = bordas
        aba[f'{colunas[k + 2]}{linha}'].font = Font(bold=True)
    
    linha = 37
    for linha_grupo, descricao in zip(tabela_geral[2], descricoes_geral[2]):
        aba[f'{colunas[1]}{linha}'] = descricao
        aba[f'{colunas[1]}{linha}'].border = bordas
        aba[f'{colunas[14]}{linha}'] = f'=SUM({colunas[2]}{linha}:{colunas[13]}{linha})'
        aba[f'{colunas[14]}{linha}'].font = Font(bold=True)
        aba[f'{colunas[14]}{linha}'].border = bordas
        for k in range(12):
            aba[f'{colunas[k + 2]}{linha}'] = linha_grupo[k]
            aba[f'{colunas[k + 2]}{linha}'].border = bordas
        linha += 1  
    aba[f'{colunas[1]}{linha}'] = 'Subtotal Componente'
    aba[f'{colunas[1]}{linha}'].border = bordas
    for k in range(13):
        aba[f'{colunas[k + 2]}{linha}'] = f'=SUM({colunas[k + 2]}37:{colunas[k + 2]}{linha-1})'
        aba[f'{colunas[k + 2]}{linha}'].border = bordas
        aba[f'{colunas[k + 2]}{linha}'].font = Font(bold=True)
    
    linha = 43
    for linha_grupo in tabela_geral[3]:
        for k in range(len(linha_grupo)):
            aba[f'{colunas[k + 2]}{linha}'] = linha_grupo[k]
            aba[f'{colunas[k + 2]}{linha}'].border = bordas
        linha += 1
    aba[f'{colunas[1]}{linha}'] = 'TOTAL GERAL'
    aba[f'{colunas[1]}{linha}'].font = Font(bold=True)
    aba[f'{colunas[1]}{linha}'].border = bordas
    if len(tabela_geral) != 0:
        aba[f'{colunas[7]}{linha}'] = f'=SUM({colunas[7]}43:{colunas[k + 2]}{linha-1})'
        aba[f'{colunas[7]}{linha}'].border = bordas
        aba[f'{colunas[7]}{linha}'].font = Font(bold=True)
        aba[f'{colunas[9]}{linha}'] = f'=SUM({colunas[9]}43:{colunas[k + 2]}{linha-1})'
        aba[f'{colunas[9]}{linha}'].border = bordas
        aba[f'{colunas[9]}{linha}'].font = Font(bold=True)
    
while True:
    try:
        planilha.save(f'Planilha de acompanhamento {ano}.xlsx')
        print('=' * 15)
        print('Arquivo salvo com SUCESSO!')
        print('=' * 15)
        break
    except PermissionError:
        print('=' * 15)
        print('Feche o arquivo para salva-lo com as alterações')
        print('=' * 15)
        sleep(5)
