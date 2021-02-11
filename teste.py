from openpyxl import Workbook, load_workbook
from openpyxl.styles.borders import Border, Side, BORDER_THIN

municipios = (
    {'nome': 'Anajás', 'cpfCnpjUg': 13715424000184, 'municipio': 150070},
    {'nome': 'Bagre', 'cpfCnpjUg': 13888332000104, 'municipio': 150110},
    {'nome': 'Breves', 'cpfCnpjUg': 17298800000133, 'municipio': 150180},
    {'nome': 'Curralinho', 'cpfCnpjUg': 11441240000148, 'municipio': 150280},
    {'nome': 'Gurupá', 'cpfCnpjUg': 12049775000130, 'municipio': 150310},
    {'nome': 'Melgaço', 'cpfCnpjUg': 11530230000189, 'municipio': 150450},
    {'nome': 'Portel', 'cpfCnpjUg': 11956268000118, 'municipio': 150580}
)

thin_border = Border(
    left=Side(border_style=BORDER_THIN, color='00000000'),
    right=Side(border_style=BORDER_THIN, color='00000000'),
    top=Side(border_style=BORDER_THIN, color='00000000'),
    bottom=Side(border_style=BORDER_THIN, color='00000000')
)



planilha = load_workbook(filename='modelo/modelo.xlsx')

for municipio in municipios:
    
    aba = planilha[municipio['nome']]
    aba.cell(row=3, column=2, value='sasldasdasdas').border = thin_border
planilha.save('teste.xlsx')